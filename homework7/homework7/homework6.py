import openpyxl as op
import numpy as np
import LinearRegression as lr

def mprtx(filename: str) -> list :
    """imports filename.xlsx and returns X [[]]"""
    wb = op.load_workbook(filename+'.xlsx')
    ws = wb.active
    x=[]
    row=2
    while True:
        if ws.cell(row,1).value == None:
            break
        tmp = [ws.cell(row,1).value,ws.cell(row,2).value]
        x.append(tmp)
        row += 1
    return x

def mprty(filename: str) -> list :
    """imports filename.xlsx and returns Y []"""
    wb = op.load_workbook(filename+'.xlsx')
    ws = wb.active
    y=[]
    row=2
    while True:
        if ws.cell(row,1).value == None:
            break
        y.append(ws.cell(row,3).value)
        row += 1
    return y

def linear_regression_decay(x: list, y: list, lmda: float) -> np.array :
    """returns linear regression with added term lmda/N sigma w"""
    X=np.array(x)
    Y=np.array(y)
    d=len(x[0])
    return np.linalg.inv(np.transpose(X)@X + lmda*np.identity(d))@np.transpose(X)@Y

def non_lin_transf(x: list) -> list :
    """gets lists of [x1,x2] and returns lists of [1,x1,x2,x1^2,x2^2,x1x2,|x1-x2|,|x1+x2|]"""
    transx=[]
    for n in range(len(x)):
        x1=x[n][0]
        x2=x[n][1]
        tmp=[1,x1,x2,x1**2,x2**2,x1*x2,abs(x1-x2),abs(x1+x2)]
        transx.append(tmp)
    return transx

def question2() :
    x=mprtx('in')
    xtr=non_lin_transf(x)
    y=mprty('in')
    xout=mprtx('out')
    xouttr=non_lin_transf(xout)
    yout=mprty('out')
    g = lr.linear_regression(xtr,y)
    Ein=lr.evaluate(xtr,y,g)
    Eout=lr.evaluate(xouttr,yout,g)
    print(1-Ein,1-Eout)

def question3() :
    x=mprtx('in')
    xtr=non_lin_transf(x)
    y=mprty('in')
    xout=mprtx('out')
    xouttr=non_lin_transf(xout)
    yout=mprty('out')
    lmda=10**(int(input('lambda= 10^k, k: ')))
    g=linear_regression_decay(xtr,y,lmda)
    Ein=lr.evaluate(xtr,y,g)
    Eout=lr.evaluate(xouttr,yout,g)
    print('Ein:',1-Ein,'Eout:',1-Eout)
